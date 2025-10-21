#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SHEETS = ["Ejercicio 2", "Ejercicio 4", "Ejercicio A", "Ejercicio B", "Ejercicio C"]


def style_header(ws, row=5):
    headers = ["Paso", "Concepto", "Cálculo/Explicación", "Resultado"]
    cols = ["A", "B", "C", "D"]
    for col, text in zip(cols, headers):
        cell = ws[f"{col}{row}"]
        cell.value = text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")

    thin = Side(style="thin", color="FFAAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in cols:
        ws[f"{col}{row}"].border = border


def style_table(ws, start_row=6, end_row=60):
    thin = Side(style="thin", color="FFDDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            if c in (2, 3):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
            cell.border = border


def set_column_widths(ws):
    widths = [10, 30, 60, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def configure_print(ws):
    # Diseño apaisado y ajuste a 1 página de ancho
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # sin limitar en alto
    # Ajuste de márgenes razonables
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    # Área de impresión sugerida
    ws.print_area = "A1:D60"
    # Repetir filas de título en cada página
    ws.print_title_rows = "1:5"


def build_cover_sheet(ws):
    ws.title = "Portada"
    ws["A1"] = "Práctica P1 — Contabilidad de Gestión"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Completa tus datos:"
    ws["A5"] = "Nombre y apellidos:"
    ws["A6"] = "DNI:"
    ws["A7"] = "Grupo:"
    ws["A8"] = "Fecha:"
    ws["A10"] = "Indicaciones:"
    ws["A11"] = "- No cambies el nombre de las hojas de los ejercicios."
    ws["A12"] = "- Si exportas a PDF, asegúrate de que las tablas no queden partidas."
    ws["A13"] = "- Un solo archivo para todos los ejercicios."

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60


def build_exercise_sheet(ws, title):
    ws.title = title
    # Título e instrucciones
    ws["A1"] = f"{title} - Respuestas"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    ws["A3"] = "Instrucciones: complete sus respuestas en la tabla. No modifique el nombre de esta hoja."
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A3:D3")

    # Encabezados y tabla
    set_column_widths(ws)
    style_header(ws, row=5)
    style_table(ws, start_row=6, end_row=60)

    # Congelar encabezados
    ws.freeze_panes = "A6"

    # Configuración de impresión
    configure_print(ws)


def generate_workbook(output_path: str):
    wb = Workbook()
    cover = wb.active
    build_cover_sheet(cover)

    for sheet_name in SHEETS:
        ws = wb.create_sheet()
        build_exercise_sheet(ws, sheet_name)

    # Asegurar que la portada esté al principio
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(cover)))

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Genera la plantilla Excel para la Práctica P1.")
    parser.add_argument("-o", "--output", default="P1_Ejercicios.xlsx", help="Ruta de salida del archivo Excel.")
    args = parser.parse_args()
    generate_workbook(args.output)
    print(f"Plantilla generada correctamente: {args.output}")


if __name__ == "__main__":
    main()
